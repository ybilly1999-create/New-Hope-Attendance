"""Admin routes — protected by simple password (session-based)."""
import io
import csv
from functools import wraps
from datetime import date, datetime, timedelta
from collections import defaultdict

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, session,
    current_app, send_file, jsonify, abort
)
from sqlalchemy import func

from .models import db, Member, Attendance, Visitor, Setting, SpecialDate
from .utils import determine_status, hk_now, status_label, upcoming_sunday

admin_bp = Blueprint("admin", __name__)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin.login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pw = request.form.get("password", "")
        if pw and pw == current_app.config.get("ADMIN_PASSWORD"):
            session["admin"] = True
            session.permanent = True
            flash("已登入管理員模式", "success")
            return redirect(request.args.get("next") or url_for("admin.dashboard"))
        flash("密碼錯誤", "error")
    return render_template("admin/login.html")


@admin_bp.route("/logout")
def logout():
    session.pop("admin", None)
    flash("已登出", "info")
    return redirect(url_for("main.index"))


# ==================== Dashboard ====================
@admin_bp.route("/")
@login_required
def dashboard():
    today = date.today()
    target_date = today if today.weekday() == 6 else upcoming_sunday(today)

    # Stats
    total_members = Member.query.filter_by(is_active=True).count()
    today_records = Attendance.query.filter_by(service_date=target_date).all()
    today_visitors = Visitor.query.filter_by(service_date=target_date).count()

    by_status = defaultdict(int)
    adult_on, adult_late, child_on, child_late = 0, 0, 0, 0
    for r in today_records:
        if r.member.is_child:
            if r.status == "on_time":
                child_on += 1
            else:
                child_late += 1
        else:
            if r.status == "on_time":
                adult_on += 1
            else:
                adult_late += 1
        by_status[r.status] += 1

    # Last 8 Sundays totals for quick overview
    last_dates = (
        db.session.query(Attendance.service_date, func.count(Attendance.id))
        .group_by(Attendance.service_date)
        .order_by(Attendance.service_date.desc())
        .limit(8)
        .all()
    )

    return render_template(
        "admin/dashboard.html",
        target_date=target_date,
        total_members=total_members,
        adult_on=adult_on, adult_late=adult_late,
        child_on=child_on, child_late=child_late,
        today_total=len(today_records),
        today_visitors=today_visitors,
        last_dates=last_dates,
    )


# ==================== Members CRUD ====================
@admin_bp.route("/members")
@login_required
def members_list():
    q = request.args.get("q", "").strip()
    show_inactive = request.args.get("inactive") == "1"
    query = Member.query
    if not show_inactive:
        query = query.filter_by(is_active=True)
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Member.name.ilike(like),
                Member.english_name.ilike(like),
                Member.member_no.ilike(like),
            )
        )
    members = query.order_by(Member.member_no).all()
    return render_template("admin/members_list.html", members=members, q=q, show_inactive=show_inactive)


@admin_bp.route("/members/new", methods=["GET", "POST"])
@login_required
def member_new():
    if request.method == "POST":
        member_no = request.form.get("member_no", "").strip().zfill(3)
        name = request.form.get("name", "").strip()
        english_name = request.form.get("english_name", "").strip() or None
        note = request.form.get("note", "").strip() or None
        is_child = request.form.get("is_child") == "1"
        if not member_no or not name:
            flash("請填寫編號與姓名", "error")
            return redirect(url_for("admin.member_new"))
        if Member.query.filter_by(member_no=member_no).first():
            flash(f"編號 {member_no} 已存在", "error")
            return redirect(url_for("admin.member_new"))
        m = Member(
            member_no=member_no, name=name, english_name=english_name,
            note=note, is_child=is_child, is_active=True,
        )
        db.session.add(m)
        db.session.commit()
        flash(f"已新增會員 {m.display_name}", "success")
        return redirect(url_for("admin.members_list"))
    # Suggest next member_no
    last = Member.query.order_by(Member.member_no.desc()).first()
    suggest = ""
    if last and last.member_no.isdigit():
        suggest = str(int(last.member_no) + 1).zfill(3)
    return render_template("admin/member_form.html", member=None, suggest=suggest)


@admin_bp.route("/members/<int:mid>/edit", methods=["GET", "POST"])
@login_required
def member_edit(mid):
    m = db.session.get(Member, mid) or abort(404)
    if request.method == "POST":
        m.member_no = request.form.get("member_no", m.member_no).strip().zfill(3)
        m.name = request.form.get("name", m.name).strip()
        m.english_name = request.form.get("english_name", "").strip() or None
        m.note = request.form.get("note", "").strip() or None
        m.is_child = request.form.get("is_child") == "1"
        m.is_active = request.form.get("is_active") == "1"
        db.session.commit()
        flash("已更新會員資料", "success")
        return redirect(url_for("admin.members_list"))
    return render_template("admin/member_form.html", member=m, suggest="")


@admin_bp.route("/members/<int:mid>/delete", methods=["POST"])
@login_required
def member_delete(mid):
    m = db.session.get(Member, mid) or abort(404)
    # Soft delete by default
    if request.form.get("hard") == "1":
        db.session.delete(m)
        flash(f"已永久刪除 {m.name}", "info")
    else:
        m.is_active = False
        flash(f"已停用 {m.name}", "info")
    db.session.commit()
    return redirect(url_for("admin.members_list"))


# ==================== Attendance management ====================
@admin_bp.route("/attendance")
@login_required
def attendance_list():
    d_str = request.args.get("date") or date.today().isoformat()
    try:
        service_d = datetime.strptime(d_str, "%Y-%m-%d").date()
    except ValueError:
        service_d = date.today()

    records = (
        db.session.query(Attendance)
        .join(Member)
        .filter(Attendance.service_date == service_d)
        .order_by(Member.member_no)
        .all()
    )
    visitors = Visitor.query.filter_by(service_date=service_d).all()

    # All members for "manual add" select
    all_members = Member.query.filter_by(is_active=True).order_by(Member.member_no).all()
    checked_ids = {r.member_id for r in records}
    not_checked = [m for m in all_members if m.id not in checked_ids]

    return render_template(
        "admin/attendance_list.html",
        records=records,
        visitors=visitors,
        service_date=service_d,
        not_checked=not_checked,
    )


@admin_bp.route("/attendance/add", methods=["POST"])
@login_required
def attendance_add():
    member_id = request.form.get("member_id", type=int)
    d_str = request.form.get("service_date")
    status = request.form.get("status", "on_time")
    try:
        service_d = datetime.strptime(d_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        flash("日期格式錯誤", "error")
        return redirect(url_for("admin.attendance_list"))
    member = db.session.get(Member, member_id)
    if not member:
        flash("找不到會員", "error")
        return redirect(url_for("admin.attendance_list", date=service_d.isoformat()))

    existing = Attendance.query.filter_by(member_id=member.id, service_date=service_d).first()
    if existing:
        existing.status = status
        existing.method = "manual"
    else:
        db.session.add(Attendance(
            member_id=member.id, service_date=service_d,
            status=status, check_in_time=hk_now(), method="manual",
        ))
    db.session.commit()
    flash(f"已記錄 {member.name}（{status_label(status)}）", "success")
    return redirect(url_for("admin.attendance_list", date=service_d.isoformat()))


@admin_bp.route("/attendance/<int:rid>/update", methods=["POST"])
@login_required
def attendance_update(rid):
    r = db.session.get(Attendance, rid) or abort(404)
    new_status = request.form.get("status")
    if new_status in ("on_time", "late"):
        r.status = new_status
        db.session.commit()
        flash("已更新狀態", "success")
    return redirect(url_for("admin.attendance_list", date=r.service_date.isoformat()))


@admin_bp.route("/attendance/<int:rid>/delete", methods=["POST"])
@login_required
def attendance_delete(rid):
    r = db.session.get(Attendance, rid) or abort(404)
    d = r.service_date
    db.session.delete(r)
    db.session.commit()
    flash("已刪除記錄", "info")
    return redirect(url_for("admin.attendance_list", date=d.isoformat()))


# ==================== Visitors ====================
@admin_bp.route("/visitors")
@login_required
def visitors_list():
    d_str = request.args.get("date")
    query = Visitor.query
    if d_str:
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date()
            query = query.filter_by(service_date=d)
        except ValueError:
            pass
    visitors = query.order_by(Visitor.service_date.desc(), Visitor.created_at.desc()).all()
    return render_template("admin/visitors_list.html", visitors=visitors, d=d_str or "")


@admin_bp.route("/visitors/<int:vid>/delete", methods=["POST"])
@login_required
def visitor_delete(vid):
    v = db.session.get(Visitor, vid) or abort(404)
    db.session.delete(v)
    db.session.commit()
    flash("已刪除訪客記錄", "info")
    return redirect(url_for("admin.visitors_list"))


# ==================== Yearly report ====================
@admin_bp.route("/report")
@login_required
def report():
    """Year summary: each member's attendance across all Sundays of the chosen year."""
    year = request.args.get("year", type=int) or date.today().year

    # All Sundays in the year
    sundays = []
    d = date(year, 1, 1)
    while d.weekday() != 6:
        d += timedelta(days=1)
    while d.year == year:
        sundays.append(d)
        d += timedelta(days=7)

    members = Member.query.filter_by(is_active=True).order_by(Member.member_no).all()
    # Build attendance map: { (member_id, date): status }
    rows = (
        db.session.query(Attendance.member_id, Attendance.service_date, Attendance.status)
        .filter(Attendance.service_date >= sundays[0], Attendance.service_date <= sundays[-1])
        .all()
    )
    att_map = {(mid, d): s for (mid, d, s) in rows}

    # Per-member summary
    summary = []
    for m in members:
        on = 0
        late = 0
        for s in sundays:
            st = att_map.get((m.id, s))
            if st == "on_time":
                on += 1
            elif st == "late":
                late += 1
        summary.append({"member": m, "on_time": on, "late": late, "total": on + late})

    # Per-Sunday totals (adult/child × on_time/late)
    weekly = []
    for s in sundays:
        a_on = a_late = c_on = c_late = 0
        for m in members:
            st = att_map.get((m.id, s))
            if not st:
                continue
            if m.is_child:
                if st == "on_time":
                    c_on += 1
                else:
                    c_late += 1
            else:
                if st == "on_time":
                    a_on += 1
                else:
                    a_late += 1
        visitors_n = Visitor.query.filter_by(service_date=s).count()
        weekly.append({
            "date": s,
            "adult_on": a_on, "adult_late": a_late,
            "child_on": c_on, "child_late": c_late,
            "visitors": visitors_n,
            "total": a_on + a_late + c_on + c_late + visitors_n,
        })

    return render_template(
        "admin/report.html",
        year=year, sundays=sundays, summary=summary, weekly=weekly,
        att_map=att_map,
    )


@admin_bp.route("/report/export")
@login_required
def report_export():
    """Export year report as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    year = request.args.get("year", type=int) or date.today().year
    sundays = []
    d = date(year, 1, 1)
    while d.weekday() != 6:
        d += timedelta(days=1)
    while d.year == year:
        sundays.append(d)
        d += timedelta(days=7)

    members = Member.query.filter_by(is_active=True).order_by(Member.member_no).all()
    rows = (
        db.session.query(Attendance.member_id, Attendance.service_date, Attendance.status)
        .filter(Attendance.service_date >= sundays[0], Attendance.service_date <= sundays[-1])
        .all()
    )
    att_map = {(mid, d): s for (mid, d, s) in rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year} 全年總表"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D4E6F1")

    headers = ["編號", "姓名", "備註", "全年次數"] + [s.strftime("%m/%d") for s in sundays]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for m in members:
        line = [m.member_no, m.name, m.note or "", 0]
        total = 0
        for s in sundays:
            st = att_map.get((m.id, s))
            if st == "on_time":
                line.append("1" if not m.is_child else "3")
                total += 1
            elif st == "late":
                line.append("2" if not m.is_child else "4")
                total += 1
            else:
                line.append("")
        line[3] = total
        ws.append(line)

    # Weekly totals at bottom
    ws.append([])
    ws.append(["", "", "成人準時"] + ["", ""] + [
        sum(1 for m in members if not m.is_child and att_map.get((m.id, s)) == "on_time")
        for s in sundays
    ])
    ws.append(["", "", "成人遲到"] + ["", ""] + [
        sum(1 for m in members if not m.is_child and att_map.get((m.id, s)) == "late")
        for s in sundays
    ])
    ws.append(["", "", "兒童準時"] + ["", ""] + [
        sum(1 for m in members if m.is_child and att_map.get((m.id, s)) == "on_time")
        for s in sundays
    ])
    ws.append(["", "", "兒童遲到"] + ["", ""] + [
        sum(1 for m in members if m.is_child and att_map.get((m.id, s)) == "late")
        for s in sundays
    ])

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    for i in range(5, 5 + len(sundays)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 7

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"chongbai_chuxi_{year}.xlsx",
        as_attachment=True,
    )


# ==================== Settings ====================
@admin_bp.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "POST":
        sst = request.form.get("service_start_time", "10:00").strip() or "10:00"
        # Validate HH:MM
        try:
            h, m = sst.split(":")
            int(h); int(m)
            Setting.set("service_start_time", sst)
            db.session.commit()
            flash("已儲存設定", "success")
        except Exception:
            flash("時間格式應為 HH:MM (例 10:00)", "error")
        return redirect(url_for("admin.settings"))
    return render_template(
        "admin/settings.html",
        service_start_time=Setting.get("service_start_time", "10:00"),
    )


# ==================== Special Dates ====================
@admin_bp.route("/special-dates")
@login_required
def special_dates_list():
    items = SpecialDate.query.order_by(SpecialDate.service_date.desc()).all()
    return render_template("admin/special_dates.html", items=items)


@admin_bp.route("/special-dates/new", methods=["POST"])
@login_required
def special_dates_new():
    d_str = request.form.get("service_date", "").strip()
    label = request.form.get("label", "").strip()
    note = request.form.get("note", "").strip()
    if not d_str or not label:
        flash("請填寫日期與名稱", "error")
        return redirect(url_for("admin.special_dates_list"))
    try:
        sd = datetime.strptime(d_str, "%Y-%m-%d").date()
    except ValueError:
        flash("日期格式錯誤", "error")
        return redirect(url_for("admin.special_dates_list"))
    existing = SpecialDate.query.filter_by(service_date=sd).first()
    if existing:
        existing.label = label
        existing.note = note or None
    else:
        db.session.add(SpecialDate(service_date=sd, label=label, note=note or None))
    db.session.commit()
    flash("已儲存特殊日期", "success")
    return redirect(url_for("admin.special_dates_list"))


@admin_bp.route("/special-dates/<int:sid>/edit", methods=["POST"])
@login_required
def special_dates_edit(sid):
    item = db.session.get(SpecialDate, sid)
    if not item:
        abort(404)
    item.label = request.form.get("label", "").strip() or item.label
    item.note = request.form.get("note", "").strip() or None
    db.session.commit()
    flash("已更新", "success")
    return redirect(url_for("admin.special_dates_list"))


@admin_bp.route("/special-dates/<int:sid>/delete", methods=["POST"])
@login_required
def special_dates_delete(sid):
    item = db.session.get(SpecialDate, sid)
    if item:
        db.session.delete(item)
        db.session.commit()
        flash("已刪除", "success")
    return redirect(url_for("admin.special_dates_list"))


# ==================== QR codes ====================
@admin_bp.route("/qr-codes")
@login_required
def qr_codes():
    """Display QR codes for printing — each member's check-in URL."""
    members = Member.query.filter_by(is_active=True).order_by(Member.member_no).all()
    base_url = request.host_url.rstrip("/")
    return render_template("admin/qr_codes.html", members=members, base_url=base_url)


@admin_bp.route("/qr-image/<member_no>.png")
@login_required
def qr_image(member_no):
    """Generate QR code PNG for a member."""
    import qrcode
    from io import BytesIO
    member_no = member_no.zfill(3)
    base_url = request.host_url.rstrip("/")
    url = f"{base_url}/qr/{member_no}"
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")

"""
Initialize database and optionally seed with member list from Excel.

Usage:
  python scripts/init_db.py                  # Just create tables
  python scripts/init_db.py --seed members.xlsx  # Seed members
  python scripts/init_db.py --seed-all year_summary.xlsx  # Seed members + historical attendance
"""
import os
import sys
import argparse
from datetime import datetime, date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app import create_app
from app.models import db, Member, Attendance, Setting, SpecialDate, Visitor


def seed_members_from_dianming(xlsx_path):
    """Seed member list from 點名用 sheet (Shou-Dong-Chong-Bai-Dian-Ming.xlsx)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "點名用" not in wb.sheetnames:
        print(f"⚠ Sheet '點名用' not found in {xlsx_path}")
        return 0
    ws = wb["點名用"]

    added = 0
    seen = set()
    for row in ws.iter_rows(values_only=True):
        for offset in [0, 4, 8, 12]:
            if offset + 2 >= len(row):
                continue
            rid, name, eng = row[offset], row[offset + 1], row[offset + 2]
            if rid is None or name is None or not isinstance(name, str):
                continue
            if isinstance(rid, float):
                rid = str(int(rid)).zfill(3)
            else:
                rid = str(rid).strip()
                if rid.isdigit():
                    rid = rid.zfill(3)
            if rid in seen:
                continue
            seen.add(rid)

            # Skip if exists
            if Member.query.filter_by(member_no=rid).first():
                continue

            # Detect children by english_name == '兒童' or note text
            is_child = False
            note = None
            eng_str = eng.strip() if isinstance(eng, str) else None
            if eng_str == "兒童":
                is_child = True
                eng_str = None
                note = "兒童"
            elif eng_str:
                note = None

            m = Member(
                member_no=rid,
                name=name.strip(),
                english_name=eng_str,
                note=note,
                is_child=is_child,
                is_active=True,
            )
            db.session.add(m)
            added += 1
    db.session.commit()
    print(f"✓ Imported {added} members from {xlsx_path}")
    return added


# Sheets that contain a yearly attendance grid (member_no | name | note | ... | dates)
# Format observed in source workbook: 53/54 sheets sharing a near-identical layout.
YEAR_SHEETS = [
    "2017記錄",
    "2018記錄",
    "2019記錄",
    "2020記錄",
    "2021記錄",
    "2022記錄",
    "「全年總表2025」",
    "全年總表",  # current year (2026)
]


def _import_year_sheet(ws, expected_year=None):
    """Import a single year sheet. Returns (new_members, new_attendance)."""
    # 1. Find the header row containing 編號 / 姓名 (usually row 7)
    header_row = None
    for r in range(1, 15):
        if ws.cell(r, 1).value == "編號" and ws.cell(r, 2).value == "姓名":
            header_row = r
            break
    if not header_row:
        print(f"  ⚠ Could not find header row in '{ws.title}'")
        return 0, 0

    # 2. Date columns: scan row 1 for datetime cells. Filter to expected year if given.
    dates = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, datetime):
            d = v.date()
            if expected_year and d.year != expected_year:
                # Skip stray cells (e.g. year-end overlap into next year is OK)
                # Only allow exact-match year, plus first week of next year
                if not (d.year == expected_year + 1 and d.month == 1 and d.day <= 7):
                    continue
            dates.append((c, d))
    if not dates:
        print(f"  ⚠ No date columns in '{ws.title}'")
        return 0, 0

    new_m = 0
    new_a = 0
    for r in range(header_row + 1, ws.max_row + 1):
        rid_raw = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        note = ws.cell(r, 3).value

        if rid_raw is None or not isinstance(name, str) or not name.strip():
            continue

        # Normalize member_no: prefer 3-digit zero-pad if it's an integer
        if isinstance(rid_raw, float):
            rid = str(int(rid_raw)).zfill(3)
        elif isinstance(rid_raw, int):
            rid = str(rid_raw).zfill(3)
        else:
            rid = str(rid_raw).strip()
            if rid.isdigit():
                rid = rid.zfill(3)

        name = name.strip()
        note_str = note.strip() if isinstance(note, str) else None

        m = Member.query.filter_by(member_no=rid).first()
        if not m:
            is_child = bool(note_str and "兒童" in note_str)
            m = Member(
                member_no=rid, name=name, note=note_str,
                is_child=is_child, is_active=True,
            )
            db.session.add(m)
            db.session.flush()
            new_m += 1
        else:
            if not m.note and note_str:
                m.note = note_str

        for col_idx, sd in dates:
            val = ws.cell(r, col_idx).value
            if val in (None, "", 0, 0.0):
                continue
            try:
                v = int(float(val))
            except (ValueError, TypeError):
                # Non-numeric (e.g. special-date label characters) — ignore
                continue
            if v not in (1, 2, 3, 4):
                continue
            if v in (3, 4):
                m.is_child = True
            status = "on_time" if v in (1, 3) else "late"
            existing = Attendance.query.filter_by(member_id=m.id, service_date=sd).first()
            if existing:
                continue
            db.session.add(Attendance(
                member_id=m.id, service_date=sd, status=status,
                check_in_time=datetime.combine(sd, datetime.min.time()),
                method="import",
            ))
            new_a += 1
    db.session.commit()
    return new_m, new_a


def seed_from_year_summary(xlsx_path, sheet_name=None):
    """Seed members AND attendance from ALL year sheets in the workbook.

    If sheet_name is given, only that sheet is imported (legacy behavior).
    Otherwise, every known year sheet is imported in chronological order.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet_name:
        targets = [sheet_name] if sheet_name in wb.sheetnames else []
    else:
        targets = [s for s in YEAR_SHEETS if s in wb.sheetnames]

    if not targets:
        print(f"⚠ No year sheets found in {xlsx_path}")
        return

    print(f"Importing {len(targets)} year sheet(s) from {xlsx_path}")
    total_m = 0
    total_a = 0
    for sn in targets:
        ws = wb[sn]
        # Extract expected year from sheet name (digits like 2017, 2018, 2025)
        import re
        m_year = re.search(r"(20\d{2})", sn)
        expected_year = int(m_year.group(1)) if m_year else None
        if sn == "全年總表":
            # Current year — derive from first date cell
            for c in range(1, ws.max_column + 1):
                v = ws.cell(1, c).value
                if isinstance(v, datetime):
                    expected_year = v.date().year
                    break
        new_m, new_a = _import_year_sheet(ws, expected_year)
        print(f"  ✓ {sn} (year={expected_year}): +{new_m} members, +{new_a} attendance")
        total_m += new_m
        total_a += new_a

    print(f"✓ Imported {total_m} new members, {total_a} attendance records from {xlsx_path}")


def _is_label_char(v):
    """Single character that could be part of a special-date label."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if len(s) != 1:
        return False
    if s.isdigit():
        return False
    if s in "-=":
        return False
    if 'A' <= s <= 'Z' or 'a' <= s <= 'z':
        return True
    return '\u4e00' <= s <= '\u9fff'


def _is_short_label(v):
    """Multi-char label already inside one cell (e.g. 會慶, 打風)."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s or s.replace(".", "").isdigit():
        return False
    if s in ("-", "="):
        return False
    if 2 <= len(s) <= 8:
        return any('\u4e00' <= ch <= '\u9fff' for ch in s)
    return False


def seed_special_dates(xlsx_path):
    """Detect special-service labels in year sheets (e.g. 會慶, BB立願禮, 夏令會,
    網上聯堂崇拜, 戶外崇拜, 打風暫停).

    Layout 1 — vertical: characters are stacked one-per-row inside a single date column,
    e.g. col 8 of the 2025 sheet has 基/督/少/年/軍/立 in rows 8-13.

    Layout 2 — horizontal-stacked: in covid years (2020-2022) one character per row
    spans across many date columns, with empty/data rows between, e.g. row 14 has
    網/網/網/... across columns, row 17 has 上/上/..., row 19 has 聯/聯/...

    Both reduce to: per date column, gather label chars from member rows, then group
    runs of chars that are vertically close (gap up to 10 rows) into one event label.
    """
    from collections import defaultdict
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    targets = [s for s in YEAR_SHEETS if s in wb.sheetnames]
    if not targets:
        return

    date_labels: dict = defaultdict(set)

    for sn in targets:
        ws = wb[sn]
        header_row = None
        for r in range(1, 20):
            if ws.cell(r, 1).value == "編號":
                header_row = r
                break
        if not header_row:
            continue
        date_cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if isinstance(v, datetime):
                date_cols.append((c, v.date()))
        if not date_cols:
            continue
        member_rows = list(range(header_row + 1, ws.max_row + 1))

        for c, d in date_cols:
            chars_at = []
            for r in member_rows:
                v = ws.cell(r, c).value
                if _is_label_char(v):
                    chars_at.append((r, v.strip()))
            if chars_at:
                # Group into runs (gap <= 10 rows separates two events)
                groups = [[chars_at[0]]]
                for r, ch in chars_at[1:]:
                    if r - groups[-1][-1][0] <= 10:
                        groups[-1].append((r, ch))
                    else:
                        groups.append([(r, ch)])
                for g in groups:
                    chars = [ch for _, ch in g]
                    # Need at least 2 chars and not all the same letter
                    if len(chars) >= 2 and len(set(chars)) > 1:
                        date_labels[d].add("".join(chars))

            # Also pick up multi-char short labels stored in a single cell
            for r in member_rows:
                v = ws.cell(r, c).value
                if _is_short_label(v):
                    date_labels[d].add(v.strip())

    # For each date, drop labels that are substrings of longer ones, and join
    # remaining distinct labels with ' / ' so admin can see all detected events.
    added = 0
    for d in sorted(date_labels.keys()):
        labels = list(date_labels[d])
        # Drop pure-repeat strings (e.g. "網網網網")
        labels = [l for l in labels if len(set(l)) > 1]
        # Drop substrings
        kept = [l for l in labels if not any(l != l2 and l in l2 for l2 in labels)]
        if not kept:
            continue
        kept.sort(key=lambda s: (-len(s), s))
        label = " / ".join(kept)
        existing = SpecialDate.query.filter_by(service_date=d).first()
        if existing:
            # Update if our detection is longer/more informative
            if not existing.label or len(label) > len(existing.label):
                existing.label = label
            continue
        db.session.add(SpecialDate(service_date=d, label=label))
        added += 1
    db.session.commit()
    print(f"✓ Imported/updated special-date labels for {len(date_labels)} dates ({added} new)")


def merge_dianming_metadata(xlsx_path):
    """Apply english_name / 兒童 flag / remarks from 點名用 sheet to existing members.

    The 點名用 sheet's third sub-column was named "Eng Name" but admins use it as a
    free-form remark column too. Common patterns we now route into note instead:
      - '兒童'          → set is_child=True; note='兒童'
      - '牧師' / '傳道'  → note holds the role
      - '六太' / '六先' / 'X太' / 'Max\u2019s 媽' / etc. → note holds the relationship
      - 'Fanny丈夫'      → note holds the relationship; english_name extracts the latin part
      - Pure latin names (Money, Raymond) → stored as english_name
    """
    import openpyxl
    import re
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "點名用" not in wb.sheetnames:
        return
    ws = wb["點名用"]
    updated = 0
    moved_to_note = 0
    for row in ws.iter_rows(values_only=True):
        for offset in [0, 4, 8, 12]:
            if offset + 2 >= len(row):
                continue
            rid, name, eng = row[offset], row[offset + 1], row[offset + 2]
            if rid is None or not isinstance(name, str):
                continue
            if isinstance(rid, float):
                rid = str(int(rid)).zfill(3)
            else:
                rid = str(rid).strip().zfill(3) if str(rid).strip().isdigit() else str(rid).strip()
            m = Member.query.filter_by(member_no=rid).first()
            if not m:
                continue
            eng_str = eng.strip() if isinstance(eng, str) else None
            if not eng_str:
                continue
            # Skip if member is already deceased — we don't overwrite.
            if m.note and ("離世" in m.note or "安息" in m.note):
                continue
            changed = False
            has_chinese = any('\u4e00' <= ch <= '\u9fff' for ch in eng_str)
            if eng_str == "兒童":
                if not m.is_child:
                    m.is_child = True
                    changed = True
                if (m.note or "") != "兒童":
                    m.note = "兒童"
                    changed = True
                if m.english_name == "兒童":
                    m.english_name = None
                    changed = True
            elif has_chinese:
                # The string is a remark. Try to split into a latin chunk + chinese remark.
                latin_match = re.findall(r"[A-Za-z][A-Za-z'’.\-]*", eng_str)
                latin_part = " ".join(latin_match).strip() if latin_match else None
                # Drop common short connectors so we keep meaningful names only
                if latin_part and latin_part.lower() in ("s", "of"):
                    latin_part = None
                # Whatever remains (Chinese + punctuation) becomes the note
                remark = eng_str
                m.note = remark
                m.english_name = latin_part
                changed = True
                moved_to_note += 1
            else:
                # Pure latin / numeric — it's an English name, not a remark
                if not m.english_name:
                    m.english_name = eng_str
                    changed = True
                # If the note column was just a duplicate of the English name (a
                # quirk left over from importing the year sheets' 3rd column),
                # clear it so the member list shows a clean note column.
                if m.note and m.note.strip() == eng_str:
                    m.note = None
                    changed = True
            if changed:
                updated += 1
    db.session.commit()
    print(f"✓ Updated metadata for {updated} members ({moved_to_note} Chinese remarks moved to note)")


# Member numbers (zero-padded 3-digit) of members who passed away —
# detected from the 備註 column in 2017記錄 / 2018記錄 sheets where it reads '離世' or '安息'.
DECEASED_MEMBERS = {
    "031": "離世",
    "079": "離世",
    "111": "離世",
    "129": "離世",
    "135": "離世",
    "161": "安息",
    "176": "離世",
    "179": "離世",
    "191": "離世",
    "192": "離世",
    "194": "離世",
    "195": "離世",
    "223": "離世",
    "225": "離世",
    "226": "離世",
}


def mark_deceased_members():
    """Set is_active=False and note='離世'/'安息' for members listed in DECEASED_MEMBERS."""
    updated = 0
    for mno, reason in DECEASED_MEMBERS.items():
        m = Member.query.filter_by(member_no=mno).first()
        if not m:
            continue
        if m.is_active or (m.note or "") != reason:
            m.is_active = False
            m.note = reason
            updated += 1
    db.session.commit()
    print(f"✓ Marked {updated} deceased members as inactive")


def seed_visitors_from_xinpengyou(xlsx_path):
    """Import 2026 visitor records from the 新朋友 sheet (cols A-Q).

    Layout: row 1 has the service date in each column; rows 2..N below contain
    one visitor name per cell. We walk each column and create a Visitor row
    for every non-empty cell.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "新朋友" not in wb.sheetnames:
        return
    ws = wb["新朋友"]
    added = 0
    # Cols A-Q = 1..17 inclusive (the 2026 block per user's note)
    for c in range(1, 18):
        v = ws.cell(1, c).value
        if not isinstance(v, datetime):
            continue
        d = v.date()
        if d.year != 2026:
            continue
        for r in range(2, 25):  # row 25 contains totals; ignore footer
            cell = ws.cell(r, c).value
            if not isinstance(cell, str):
                continue
            name = cell.strip()
            if not name:
                continue
            # Skip rows that hold totals or numbers
            if name.replace(".", "").isdigit():
                continue
            existing = (
                Visitor.query.filter_by(service_date=d, name=name).first()
            )
            if existing:
                continue
            db.session.add(Visitor(service_date=d, name=name))
            added += 1
    db.session.commit()
    print(f"✓ Imported {added} 2026 visitor records from 新朋友 sheet")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed-members", help="Path to Shou-Dong-Chong-Bai-Dian-Ming.xlsx")
    parser.add_argument("--seed-history", help="Path to Chong-Bai-Chu-Xi-Ji-Lu-2025-2026.xlsx")
    parser.add_argument("--reset", action="store_true", help="Drop all tables first (DANGER)")
    args = parser.parse_args()

    app = create_app()
    with app.app_context():
        if args.reset:
            print("⚠ Resetting database...")
            db.drop_all()
            db.create_all()

        # Always ensure tables exist
        db.create_all()
        print(f"✓ Database ready: {app.config['SQLALCHEMY_DATABASE_URI']}")

        existing = Member.query.count()
        print(f"Current members in DB: {existing}")

        if args.seed_history and Path(args.seed_history).exists():
            seed_from_year_summary(args.seed_history)
            seed_special_dates(args.seed_history)

        if args.seed_members and Path(args.seed_members).exists():
            seed_members_from_dianming(args.seed_members)
            merge_dianming_metadata(args.seed_members)

        # Always run cleanup steps (idempotent)
        mark_deceased_members()
        if args.seed_history and Path(args.seed_history).exists():
            seed_visitors_from_xinpengyou(args.seed_history)

        print(
            f"Final members: {Member.query.count()}, "
            f"attendance: {Attendance.query.count()}, "
            f"visitors: {Visitor.query.count()}, "
            f"special dates: {SpecialDate.query.count()}"
        )


if __name__ == "__main__":
    main()
